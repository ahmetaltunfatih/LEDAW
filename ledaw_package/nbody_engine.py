import os
import re
import glob
import shutil
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from .classes import *


def label_systems(filenames):
    """Assign SUPERSYS to the first file and SUBSYS1, SUBSYS2, ... to the rest."""
    labeled_filenames = {}
    for i, filename in enumerate(filenames):
        if filename:  # Ensure filename is not None or empty
            if i == 0:
                labeled_filenames[filename] = "SUPERSYS"
            else:
                labeled_filenames[filename] = f"SUBSYS{i}"
    return labeled_filenames


def parse_coordinates_from_file(file_path):
    """Extract unique labels and coordinates from the given file."""
    label_coord_map = {}
    fragment_mode = False
    frag_label_counter = 1
    current_fragment = None
    pattern_found = False

    with open(file_path, 'r') as file:
        lines = file.readlines()

    for line in lines:
        if re.search(r'CARTESIAN COORDINATES OF FRAGMENTS \(ANGSTROEM\)', line):
            fragment_mode = True
            label_coord_map.clear()
            frag_label_counter = 1
            pattern_found = True
            continue
        
        if re.search(r'INTERNAL COORDINATES \(ANGSTROEM\)', line):
            if fragment_mode:
                fragment_mode = False
            continue

        if fragment_mode:
            fragment_match = re.match(r'\s*FRAGMENT\s+(\d+)', line)
            if fragment_match:
                current_fragment = int(fragment_match.group(1))
                continue
            
            coord_match = re.match(r'\s*\w+\s+([-.\d]+)\s+([-.\d]+)\s+([-.\d]+)', line)
            if coord_match and current_fragment is not None and current_fragment not in label_coord_map:
                coordinates = (float(coord_match.group(1)), float(coord_match.group(2)), float(coord_match.group(3)))
                label_coord_map[current_fragment] = coordinates
                current_fragment = None

    if not pattern_found:
        fragment_mode = False
        for line in lines:
            if re.search(r'CARTESIAN COORDINATES \(ANGSTROEM\)', line):
                fragment_mode = True
                label_coord_map.clear()
                frag_label_counter = 1
                continue
            
            if re.search(r'CARTESIAN COORDINATES \(A\.U\.\)', line):
                if fragment_mode:
                    fragment_mode = False
                continue
            
            if fragment_mode:
                coord_match = re.match(r'\s*\w+\s+([-.\d]+)\s+([-.\d]+)\s+([-.\d]+)', line)
                if coord_match:
                    label = frag_label_counter
                    frag_label_counter += 1
                    coordinates = (float(coord_match.group(1)), float(coord_match.group(2)), float(coord_match.group(3)))
                    if label not in label_coord_map:
                        label_coord_map[label] = coordinates

    return label_coord_map


def coordinates_match(coord1, coord2, tol=1e-3):
    """Check if two coordinates match within a given tolerance using numpy."""
    return np.allclose(coord1, coord2, atol=tol)

 
def match_and_construct_mappings(supersystem_coords, other_coords_list, tol=1e-3):
    """Match fragment coordinates to supersystem labels and construct the mapping lists."""
    mapped_labels_list = []
    match_dicts = []
    subsystem_matching_labels = []

    for frag_coords in other_coords_list:
        match_dict = {}
        label_swaps = {}
        matched_labels = []

        # Match supersystem labels with fragment labels and swap them
        for super_label, super_coord in supersystem_coords.items():
            found_match = False
            for frag_label, frag_coord in frag_coords.items():
                if coordinates_match(super_coord, frag_coord, tol):
                    # Store the swap and the matching labels
                    label_swaps[super_label] = frag_label
                    label_swaps[frag_label] = super_label
                    matched_labels.append(super_label)
                    found_match = True
                    break

            if not found_match:
                # Only append `super_label` if a match was found
                matched_labels.append(None)

        # Now construct the final match dictionary based on the swaps
        for i in range(1, len(supersystem_coords) + 1):
            if i in label_swaps:
                match_dict[i] = label_swaps[i]
            else:
                match_dict[i] = i

        # Create the mapped labels list from the match_dict
        mapped_labels = [match_dict[i] for i in sorted(match_dict)]
        mapped_labels_list.append(mapped_labels)
        match_dicts.append(match_dict)

        # Append only matched labels to the subsystem matching list
        matched_labels_cleaned = [label for label in matched_labels if label is not None]
        subsystem_matching_labels.append(matched_labels_cleaned)

    return mapped_labels_list, match_dicts, subsystem_matching_labels


def construct_label_mappings(main_or_alt_filenames, default_label_mappings=None):
    """Constructs label mappings for a supersystem and its subsystems based on coordinate matching."""
    if not main_or_alt_filenames[0] and default_label_mappings:
        return default_label_mappings, [{}], []

    if not main_or_alt_filenames[0]:
        raise ValueError("The first file (supersystem) in the list cannot be empty.")

    # Parse the supersystem file and assign sequential labels
    supersystem_coords = parse_coordinates_from_file(main_or_alt_filenames[0])

    # Filter out any empty or None entries in the file list before processing
    other_coords_list = [parse_coordinates_from_file(filename) for filename in main_or_alt_filenames[1:] if filename]

    # Match and construct mappings based on coordinate matching
    main_or_alt_label_mappings, match_dicts, subsystem_matching_labels = match_and_construct_mappings(supersystem_coords, other_coords_list)

    # Add the supersystem labels as the first list
    main_or_alt_label_mappings.insert(0, list(range(1, len(supersystem_coords) + 1)))

    return main_or_alt_label_mappings, match_dicts, subsystem_matching_labels


def extract_system_name(filename):
    """Extract system name from the filename considering the last slash and first dot."""
    if not filename:  # Handle None or empty string
        return ""
    
    # Find the position of the last slash or backslash
    last_slash_pos = max(filename.rfind('/'), filename.rfind('\\'))
    
    # Determine the starting position for extraction
    start_pos = last_slash_pos + 1 if last_slash_pos != -1 else 0
    
    # Find the position of the first dot after the last slash
    dot_pos = filename.find('.', start_pos)
    
    # Extract the system name
    system_name = filename[start_pos:dot_pos] if dot_pos != -1 else filename[start_pos:]
    
    return system_name


def compute_diel_int_en(main_filenames, alternative_filenames, conversion_factor):
    """Extract CPCM Dielectric values from all main and alternative files and calculate dielectric interaction energy."""
    
    diel_values = {}
    super_system_diel = None
    subsystem_diel_values = {}
    
    diel_pattern = r"CPCM Dielectric\s*:\s*([-+]?\d*\.\d+|\d+)"  
    
    # Function to extract dielectric value from a file
    def extract_diel_from_file(filename):
        """Extract the dielectric value from a given file."""
        try:
            with open(filename, 'r') as file:
                for line in file:
                    match = re.search(diel_pattern, line)
                    if match:
                        return float(match.group(1))
        except FileNotFoundError:
            print(f"File {filename} not found.")
        return None

    # Loop over each main and alternative file
    for main_filename, alt_filename in zip(main_filenames, alternative_filenames):
        root_name = extract_system_name(main_filename)
        
        # Try to extract from the main file
        diel_value = extract_diel_from_file(main_filename)
        
        # If not found, try the alternative file
        if diel_value is None and alt_filename:
            diel_value = extract_diel_from_file(alt_filename)
        
        diel_values[root_name] = diel_value
        
        # Check if this is the supersystem
        if root_name == extract_system_name(main_filenames[0]):
            super_system_diel = diel_value
        else:
            if diel_value is not None:
                subsystem_diel_values[root_name] = diel_value

    # Check if any dielectric value is missing
    if super_system_diel is None or len(subsystem_diel_values) != (len(main_filenames) - 1):
        print("At least one of the output files does not contain dielectric contribution.\nDielectric contribution to the interaction energy is taken as 0")
        return diel_values, 0.0

    # Calculate the dielectric interaction energy
    diel_int_energy = (super_system_diel - sum(subsystem_diel_values.values())) * conversion_factor
    
    return diel_values, diel_int_energy

	
def check_local_energy_decomposition(filename, patterns):
    """Check if the file contains the 'LOCAL ENERGY DECOMPOSITION' section."""
    root_name = extract_system_name(filename)
    try:
        with open(filename, 'r') as file:
            content = file.read()
        return bool(re.search(patterns.local_energy_decomp_pattern, content))
    except FileNotFoundError:
        print(f"File {filename} not found. Skipping.")
        return False


def extract_numbers(pattern, content):
    """Extracts a list of numbers based on the regex pattern from the content."""
    numbers = []
    match = re.search(pattern, content)
    if match:
        number_pattern = r"([-]?\d*\.\d+)"
        numbers = [float(number) for number in re.findall(number_pattern, match.group(0))]
    return numbers


def extract_first_match_from_file(filename, patterns, method, use_ref_as_rhf_in_hfld=None):
    """Extract the first matches for E(0), strong pairs, weak pairs, and triples correction from the file."""
    root_name = extract_system_name(filename)
    try:
        with open(filename, 'r') as file:
            content = file.read()
        
        # Search for the energy patterns
        e_ref_match = re.search(patterns.pattern_e0, content)
        e_sp_match = re.search(patterns.pattern_strong_corr, content)
        e_wp_match = re.search(patterns.pattern_weak_corr, content)
        e_t_match = re.search(patterns.pattern_triples_corr, content)

        # Extract the values, default to 0 if not found
        e_ref = float(e_ref_match.group(1)) if e_ref_match else 0.0
        e_sp = float(e_sp_match.group(1)) if e_sp_match else 0.0
        e_wp = float(e_wp_match.group(1)) if e_wp_match else 0.0
        e_t = float(e_t_match.group(1)) if e_t_match else 0.0

        # Fallback for HFLD method if E(0) is not found and use_ref_as_rhf_in_hfld is True
        if e_ref == 0.0 and method.lower() == 'hfld' and use_ref_as_rhf_in_hfld:
            total_energy_match = re.search(r"Total Energy\s+:\s+([-]?\d+\.\d+)", content)
            dielectric_match = re.search(r"CPCM Dielectric\s+:\s+([-]?\d+\.\d+)", content)
            
            if total_energy_match:
                total_energy = float(total_energy_match.group(1))
                dielectric_value = float(dielectric_match.group(1)) if dielectric_match else 0.0
                e_ref = total_energy - dielectric_value
        
        # Apply fallback mechanism for HFLD method on strong pairs, weak pairs, and triples correction
        if use_ref_as_rhf_in_hfld and method.lower() == 'hfld':
            e_sp = e_sp if e_sp != 0.0 else e_ref
            e_wp = e_wp if e_wp != 0.0 else e_ref
            e_t = e_t if e_t != 0.0 else e_ref

        return e_ref, e_sp, e_wp, e_t
    
    except FileNotFoundError:
        print(f"File {filename} not found. Skipping.")
        return 0.0, 0.0, 0.0, 0.0


def determine_matrix_size(filename, patterns):
    """Determine the size of the matrix based on the intra_ref pattern in the file."""
    root_name = extract_system_name(filename)
    try:
        with open(filename, 'r') as file:
            content = file.read()
        intra_ref = extract_numbers(patterns.PATTERNS["intra_ref"], content)
        if not intra_ref:
            intra_ref = extract_numbers(patterns.PATTERNS["intra_ref_alt"], content)
        return len(intra_ref)
    except FileNotFoundError:
        raise FileNotFoundError(f"File {filename} not found. Unable to determine matrix size.")

        
def extract_els_exch_matrices(content, primary_pattern, alternative_pattern):
    """Extracts electrostatics and exchange matrices based on the primary and alternative patterns."""
    matches = re.findall(primary_pattern, content)
    if not matches:
        matches = re.findall(alternative_pattern, content)
    if not matches:
        return None, None

    fragment_ids = set()
    for match in matches:
        fragment_ids.add(int(match[0]))
        fragment_ids.add(int(match[1]))
    matrix_size = max(fragment_ids)

    electrostatics_matrix = np.zeros((matrix_size, matrix_size))
    exchange_matrix = np.zeros((matrix_size, matrix_size))

    for match in matches:
        i = int(match[0]) - 1 
        j = int(match[1]) - 1 
        els_value = float(match[2])
        exch_value = float(match[3])
        electrostatics_matrix[i, j] = electrostatics_matrix[j, i] = els_value
        exchange_matrix[i, j] = exchange_matrix[j, i] = exch_value

    return electrostatics_matrix, exchange_matrix


def process_main_file(filename, matrix_size, patterns, intra_ref_list, intra_corr_list, intra_strong_pairs_list,
                      intra_triples_list, intra_weak_pairs_list, singles_contribution_list,
                      inter_ref_matrices, inter_corr_matrices, electrostat_matrices,
                      exchange_matrices, inter_strong_pairs_matrices, inter_triples_matrices,
                      inter_weak_pairs_matrices, dispersion_strong_pairs_matrices, system_label):
    """Process the main file to extract and store data."""
    
    try:
        with open(filename, 'r') as file:
            content = file.read()
    except FileNotFoundError:
        print(f"Main file {filename} not found. Skipping.")
        return

    intra_ref = extract_numbers(patterns.PATTERNS["intra_ref"], content)
    if not intra_ref:
        intra_ref = extract_numbers(patterns.PATTERNS["intra_ref_alt"], content)
    intra_corr = extract_numbers(patterns.PATTERNS["intra_corr"], content)
    intra_strong_pairs = extract_numbers(patterns.PATTERNS["intra_strong_pairs"], content)
    intra_triples = extract_numbers(patterns.PATTERNS["intra_triples"], content)
    intra_weak_pairs = extract_numbers(patterns.PATTERNS["intra_weak_pairs"], content)
    singles_contribution = extract_numbers(patterns.PATTERNS["singles_contribution"], content)

    # Initialize matrices with zeros
    inter_ref_matrix = np.zeros((matrix_size, matrix_size))
    inter_corr_matrix = np.zeros((matrix_size, matrix_size))
    electrostatics_matrix = np.zeros((matrix_size, matrix_size))
    exchange_matrix = np.zeros((matrix_size, matrix_size))
    inter_strong_pairs_matrix = np.zeros((matrix_size, matrix_size))
    inter_triples_matrix = np.zeros((matrix_size, matrix_size))
    inter_weak_pairs_matrix = np.zeros((matrix_size, matrix_size))
    dispersion_strong_pairs_matrix = np.zeros((matrix_size, matrix_size))

    # Find all matches
    matches_combined = re.findall(patterns.PATTERNS["ref_corr_inter"], content)
    matches_inter_correlation_full = re.findall(patterns.PATTERNS["corr_inter_full"], content)
    matches_inter_correlation_partial = re.findall(patterns.PATTERNS["corr_inter_partial"], content)
    matches_dispersion = re.findall(patterns.PATTERNS["dispersion_strong_pairs"], content)

    # Extract electrostatics and exchange data using primary and alternative patterns
    electrostat_pattern = patterns.PATTERNS["ref_inter"]
    electrostat_alternative_pattern = patterns.PATTERNS["ref_corr_inter"]
    electrostat_matrix, exch_matrix = extract_els_exch_matrices(content, electrostat_pattern, electrostat_alternative_pattern)
    if electrostat_matrix is not None:
        electrostatics_matrix = electrostat_matrix
        exchange_matrix = exch_matrix

    def populate_matrix(matches, matrix, idx):
        for match in matches:
            try:
                i = int(match[0]) - 1
                j = int(match[1]) - 1
                if i < matrix_size and j < matrix_size and idx < len(match):
                    value = float(match[idx])
                    matrix[i, j] = matrix[j, i] = value
            except (IndexError, ValueError):
                continue

    # Populate matrices
    populate_matrix(matches_combined, inter_ref_matrix, 2)
    populate_matrix(matches_combined, inter_corr_matrix, 3)
    if matches_inter_correlation_full:
        populate_matrix(matches_inter_correlation_full, inter_strong_pairs_matrix, 2)
        populate_matrix(matches_inter_correlation_full, inter_triples_matrix, 3)
        populate_matrix(matches_inter_correlation_full, inter_weak_pairs_matrix, 4)
    else:
        populate_matrix(matches_inter_correlation_partial, inter_strong_pairs_matrix, 2)
        populate_matrix(matches_inter_correlation_partial, inter_weak_pairs_matrix, 3)

    # Populate dispersion strong pairs matrix
    for match in matches_dispersion:
        try:
            i = int(match[0]) - 1
            j = int(match[1]) - 1
            value = float(match[2])
            if i < matrix_size and j < matrix_size:
                dispersion_strong_pairs_matrix[i, j] = dispersion_strong_pairs_matrix[j, i] = value
        except (IndexError, ValueError):
            continue

    # Use the system label as the file prefix
    file_prefix = system_label

    # Store results in the respective lists
    intra_ref_list.append((intra_ref, file_prefix))
    intra_corr_list.append((intra_corr, file_prefix))
    intra_strong_pairs_list.append((intra_strong_pairs, file_prefix))
    intra_triples_list.append((intra_triples, file_prefix))
    intra_weak_pairs_list.append((intra_weak_pairs, file_prefix))
    singles_contribution_list.append((singles_contribution, file_prefix))
    inter_ref_matrices.append((inter_ref_matrix, file_prefix))
    inter_corr_matrices.append((inter_corr_matrix, file_prefix))
    electrostat_matrices.append((electrostatics_matrix, file_prefix))
    exchange_matrices.append((exchange_matrix, file_prefix))
    inter_strong_pairs_matrices.append((inter_strong_pairs_matrix, file_prefix))
    inter_triples_matrices.append((inter_triples_matrix, file_prefix))
    inter_weak_pairs_matrices.append((inter_weak_pairs_matrix, file_prefix))
    dispersion_strong_pairs_matrices.append((dispersion_strong_pairs_matrix, file_prefix))


def process_alternative_file(filename, matrix_size, patterns, intra_ref_list, intra_corr_list, intra_strong_pairs_list,
                             intra_triples_list, intra_weak_pairs_list, singles_contribution_list,
                             inter_ref_matrices, inter_corr_matrices, electrostat_matrices,
                             exchange_matrices, inter_strong_pairs_matrices, inter_triples_matrices,
                             inter_weak_pairs_matrices, dispersion_strong_pairs_matrices, system_label):
    """Process the alternative file to extract and store data."""
    if not filename:
        return

    try:
        with open(filename, 'r') as file:
            content = file.read()
    except FileNotFoundError:
        print(f"Alternative file {filename} not found. Skipping.")
        return

    intra_ref = extract_numbers(patterns.PATTERNS["intra_ref"], content)
    if not intra_ref:
        intra_ref = extract_numbers(patterns.PATTERNS["intra_ref_alt"], content)
    intra_corr = extract_numbers(patterns.PATTERNS["intra_corr"], content)
    intra_strong_pairs = extract_numbers(patterns.PATTERNS["intra_strong_pairs"], content)
    intra_triples = extract_numbers(patterns.PATTERNS["intra_triples"], content)
    intra_weak_pairs = extract_numbers(patterns.PATTERNS["intra_weak_pairs"], content)
    singles_contribution = extract_numbers(patterns.PATTERNS["singles_contribution"], content)

    # Initialize matrices with zeros
    inter_ref_matrix = np.zeros((matrix_size, matrix_size))
    inter_corr_matrix = np.zeros((matrix_size, matrix_size))
    electrostatics_matrix = np.zeros((matrix_size, matrix_size))
    exchange_matrix = np.zeros((matrix_size, matrix_size))
    inter_strong_pairs_matrix = np.zeros((matrix_size, matrix_size))
    inter_triples_matrix = np.zeros((matrix_size, matrix_size))
    inter_weak_pairs_matrix = np.zeros((matrix_size, matrix_size))
    dispersion_strong_pairs_matrix = np.zeros((matrix_size, matrix_size))

    # Find all matches
    matches_combined = re.findall(patterns.PATTERNS["ref_corr_inter"], content)
    matches_inter_correlation_full = re.findall(patterns.PATTERNS["corr_inter_full"], content)
    matches_inter_correlation_partial = re.findall(patterns.PATTERNS["corr_inter_partial"], content)
    matches_dispersion = re.findall(patterns.PATTERNS["dispersion_strong_pairs"], content)

    # Extract electrostatics and exchange data using primary and alternative patterns
    electrostat_pattern = patterns.PATTERNS["ref_inter"]
    electrostat_alternative_pattern = patterns.PATTERNS["ref_corr_inter"]
    electrostat_matrix, exch_matrix = extract_els_exch_matrices(content, electrostat_pattern, electrostat_alternative_pattern)
    if electrostat_matrix is not None:
        electrostatics_matrix = electrostat_matrix
        exchange_matrix = exch_matrix

    def populate_matrix(matches, matrix, idx):
        for match in matches:
            try:
                i = int(match[0]) - 1
                j = int(match[1]) - 1
                if i < matrix_size and j < matrix_size and idx < len(match):
                    value = float(match[idx])
                    matrix[i, j] = matrix[j, i] = value
            except (IndexError, ValueError):
                continue

    # Populate matrices
    populate_matrix(matches_combined, inter_ref_matrix, 2)
    populate_matrix(matches_combined, inter_corr_matrix, 3)
    if matches_inter_correlation_full:
        populate_matrix(matches_inter_correlation_full, inter_strong_pairs_matrix, 2)
        populate_matrix(matches_inter_correlation_full, inter_triples_matrix, 3)
        populate_matrix(matches_inter_correlation_full, inter_weak_pairs_matrix, 4)
    else:
        populate_matrix(matches_inter_correlation_partial, inter_strong_pairs_matrix, 2)
        populate_matrix(matches_inter_correlation_partial, inter_weak_pairs_matrix, 3)

    # Populate dispersion strong pairs matrix
    for match in matches_dispersion:
        try:
            i = int(match[0]) - 1
            j = int(match[1]) - 1
            value = float(match[2])
            if i < matrix_size and j < matrix_size:
                dispersion_strong_pairs_matrix[i, j] = dispersion_strong_pairs_matrix[j, i] = value
        except (IndexError, ValueError):
            continue

    # Use the system label and add ' ALT' to the prefix
    file_prefix = system_label + ' ALT'

    # Store results in the respective lists
    intra_ref_list.append((intra_ref, file_prefix))
    intra_corr_list.append((intra_corr, file_prefix))
    intra_strong_pairs_list.append((intra_strong_pairs, file_prefix))
    intra_triples_list.append((intra_triples, file_prefix))
    intra_weak_pairs_list.append((intra_weak_pairs, file_prefix))
    singles_contribution_list.append((singles_contribution, file_prefix))
    inter_ref_matrices.append((inter_ref_matrix, file_prefix))
    inter_corr_matrices.append((inter_corr_matrix, file_prefix))
    electrostat_matrices.append((electrostatics_matrix, file_prefix))
    exchange_matrices.append((exchange_matrix, file_prefix))
    inter_strong_pairs_matrices.append((inter_strong_pairs_matrix, file_prefix))
    inter_triples_matrices.append((inter_triples_matrix, file_prefix))
    inter_weak_pairs_matrices.append((inter_weak_pairs_matrix, file_prefix))
    dispersion_strong_pairs_matrices.append((dispersion_strong_pairs_matrix, file_prefix))


def write_matrices_to_excel(filename, matrix_size, intra_ref_list, intra_corr_list, intra_strong_pairs_list,
                            intra_triples_list, intra_weak_pairs_list, singles_contribution_list,
                            inter_ref_matrices, inter_corr_matrices, electrostat_matrices,
                            exchange_matrices, inter_strong_pairs_matrices, inter_triples_matrices,
                            inter_weak_pairs_matrices, dispersion_strong_pairs_matrices):
    """Writes collected data to an Excel file."""
    def expand_matrix_to_size(matrix, target_size):
        """Expand a matrix to the target size with zeros."""
        current_size = matrix.shape[0]
        if current_size < target_size:
            expanded_matrix = np.zeros((target_size, target_size))
            expanded_matrix[:current_size, :current_size] = matrix
            return expanded_matrix
        return matrix

    with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
        # Write intra values as diagonal elements in separate sheets
        for idx, (intra_ref, file_prefix) in enumerate(intra_ref_list):
            file_prefix = extract_system_name(file_prefix)
            intra_ref_matrix = np.zeros((matrix_size, matrix_size))
            for i in range(len(intra_ref)):
                intra_ref_matrix[i, i] = intra_ref[i]
            df_intra_ref = pd.DataFrame(intra_ref_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df_intra_ref.to_excel(writer, sheet_name=f'Intra REF {file_prefix}')

        for idx, (intra_strong_pairs, file_prefix) in enumerate(intra_strong_pairs_list):
            file_prefix = extract_system_name(file_prefix)
            intra_strong_pairs_matrix = np.zeros((matrix_size, matrix_size))
            for i in range(len(intra_strong_pairs)):
                intra_strong_pairs_matrix[i, i] = intra_strong_pairs[i]
            df_intra_strong_pairs = pd.DataFrame(intra_strong_pairs_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df_intra_strong_pairs.to_excel(writer, sheet_name=f'Intra SP {file_prefix}')

        for idx, (intra_triples, file_prefix) in enumerate(intra_triples_list):
            file_prefix = extract_system_name(file_prefix)
            intra_triples_matrix = np.zeros((matrix_size, matrix_size))
            for i in range(len(intra_triples)):
                intra_triples_matrix[i, i] = intra_triples[i]
            df_intra_triples = pd.DataFrame(intra_triples_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df_intra_triples.to_excel(writer, sheet_name=f'Intra T {file_prefix}')

        for idx, (intra_weak_pairs, file_prefix) in enumerate(intra_weak_pairs_list):
            file_prefix = extract_system_name(file_prefix)
            intra_weak_pairs_matrix = np.zeros((matrix_size, matrix_size))
            for i in range(len(intra_weak_pairs)):
                intra_weak_pairs_matrix[i, i] = intra_weak_pairs[i]
            df_intra_weak_pairs = pd.DataFrame(intra_weak_pairs_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df_intra_weak_pairs.to_excel(writer, sheet_name=f'Intra WP {file_prefix}')

        for idx, (singles_contribution, file_prefix) in enumerate(singles_contribution_list):
            file_prefix = extract_system_name(file_prefix)
            singles_contribution_matrix = np.zeros((matrix_size, matrix_size))
            for i in range(len(singles_contribution)):
                singles_contribution_matrix[i, i] = singles_contribution[i]
            df_singles_contribution = pd.DataFrame(singles_contribution_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df_singles_contribution.to_excel(writer, sheet_name=f'Singles {file_prefix}')

        # Write interaction matrices
        for idx, (matrix, file_prefix) in enumerate(inter_ref_matrices):
            file_prefix = extract_system_name(file_prefix)
            expanded_matrix = expand_matrix_to_size(matrix, matrix_size)
            df = pd.DataFrame(expanded_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df.to_excel(writer, sheet_name=f'Inter REF {file_prefix}')

        for idx, (matrix, file_prefix) in enumerate(electrostat_matrices):
            file_prefix = extract_system_name(file_prefix)
            expanded_matrix = expand_matrix_to_size(matrix, matrix_size)
            df = pd.DataFrame(expanded_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df.to_excel(writer, sheet_name=f'Electrostat {file_prefix}')

        for idx, (matrix, file_prefix) in enumerate(exchange_matrices):
            file_prefix = extract_system_name(file_prefix)
            expanded_matrix = expand_matrix_to_size(matrix, matrix_size)
            df = pd.DataFrame(expanded_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df.to_excel(writer, sheet_name=f'Exchange {file_prefix}')

        for idx, (matrix, file_prefix) in enumerate(inter_strong_pairs_matrices):
            file_prefix = extract_system_name(file_prefix)
            expanded_matrix = expand_matrix_to_size(matrix, matrix_size)
            df = pd.DataFrame(expanded_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df.to_excel(writer, sheet_name=f'Inter SP {file_prefix}')

        for idx, (matrix, file_prefix) in enumerate(inter_triples_matrices):
            file_prefix = extract_system_name(file_prefix)
            expanded_matrix = expand_matrix_to_size(matrix, matrix_size)
            df = pd.DataFrame(expanded_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df.to_excel(writer, sheet_name=f'Inter T {file_prefix}')

        for idx, (matrix, file_prefix) in enumerate(inter_weak_pairs_matrices):
            file_prefix = extract_system_name(file_prefix)
            expanded_matrix = expand_matrix_to_size(matrix, matrix_size)
            df = pd.DataFrame(expanded_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df.to_excel(writer, sheet_name=f'Inter WP {file_prefix}')

        for idx, (matrix, file_prefix) in enumerate(dispersion_strong_pairs_matrices):
            file_prefix = extract_system_name(file_prefix)
            expanded_matrix = expand_matrix_to_size(matrix, matrix_size)
            df = pd.DataFrame(expanded_matrix, columns=range(1, matrix_size + 1), index=range(1, matrix_size + 1))
            df.to_excel(writer, sheet_name=f'Disp SP {file_prefix}')


def multifrag_system_processing(main_filenames, alternative_filenames, LEDAW_output_path, system_labels):
    # Initialize lists to store results for multiple files
    intra_ref_list = []
    intra_corr_list = []
    intra_strong_pairs_list = []
    intra_triples_list = []
    intra_weak_pairs_list = []
    singles_contribution_list = []
    inter_ref_matrices = []
    inter_corr_matrices = []
    electrostat_matrices = []
    exchange_matrices = []
    inter_strong_pairs_matrices = []
    inter_triples_matrices = []
    inter_weak_pairs_matrices = []
    dispersion_strong_pairs_matrices = []

    # Instantiate patterns object
    patterns = Patterns()

    # Determine the matrix size based on the first main file
    matrix_size = determine_matrix_size(main_filenames[0], patterns)

    # Process each main file and alternative file
    for main_file, alt_file, system_label in zip(main_filenames, alternative_filenames, system_labels):
        process_main_file(main_file, matrix_size, patterns, intra_ref_list, intra_corr_list, intra_strong_pairs_list,
                          intra_triples_list, intra_weak_pairs_list, singles_contribution_list,
                          inter_ref_matrices, inter_corr_matrices, electrostat_matrices,
                          exchange_matrices, inter_strong_pairs_matrices, inter_triples_matrices,
                          inter_weak_pairs_matrices, dispersion_strong_pairs_matrices, system_label)
        process_alternative_file(alt_file, matrix_size, patterns, intra_ref_list, intra_corr_list, intra_strong_pairs_list,
                                 intra_triples_list, intra_weak_pairs_list, singles_contribution_list,
                                 inter_ref_matrices, inter_corr_matrices, electrostat_matrices,
                                 exchange_matrices, inter_strong_pairs_matrices, inter_triples_matrices,
                                 inter_weak_pairs_matrices, dispersion_strong_pairs_matrices, system_label)

    # Ensure the output directory exists
    os.makedirs(LEDAW_output_path, exist_ok=True)

    # Define the full path for the output file
    output_file = os.path.join(LEDAW_output_path, 'tmp1.xlsx')

    # Write the collected data to an Excel file
    write_matrices_to_excel(output_file, matrix_size, intra_ref_list, intra_corr_list, intra_strong_pairs_list,
                            intra_triples_list, intra_weak_pairs_list, singles_contribution_list,
                            inter_ref_matrices, inter_corr_matrices, electrostat_matrices,
                            exchange_matrices, inter_strong_pairs_matrices, inter_triples_matrices,
                            inter_weak_pairs_matrices, dispersion_strong_pairs_matrices)

    return matrix_size


def singlefrag_system_processing(main_filenames, alternative_filenames, matrix_size, LEDAW_output_path, system_labels, method, use_ref_as_rhf_in_hfld=None):
    patterns = Patterns()
    matrices = {}

    for i, (main_filename, alt_filename, system_label) in enumerate(zip(main_filenames, alternative_filenames, system_labels)):
        # Process main file
        if main_filename:
            led_exists = check_local_energy_decomposition(main_filename, patterns)
            if not led_exists:
                e_ref, e_sp, e_wp, e_t = extract_first_match_from_file(main_filename, patterns, method, use_ref_as_rhf_in_hfld)
                ref_matrix = np.zeros((matrix_size, matrix_size))
                sp_matrix = np.zeros((matrix_size, matrix_size))
                wp_matrix = np.zeros((matrix_size, matrix_size))
                t_matrix = np.zeros((matrix_size, matrix_size))

                # Place values on the first diagonal element
                ref_matrix[0, 0] = e_ref
                sp_matrix[0, 0] = e_sp
                wp_matrix[0, 0] = e_wp
                t_matrix[0, 0] = e_t

                # Use the system label as the prefix instead of the file name
                file_prefix = system_label
                matrices[f'Intra REF {file_prefix}'] = ref_matrix
                matrices[f'Intra SP {file_prefix}'] = sp_matrix
                matrices[f'Intra WP {file_prefix}'] = wp_matrix
                matrices[f'Intra T {file_prefix}'] = t_matrix

        # Process ALT file similarly if needed
        if alt_filename:
            led_exists = check_local_energy_decomposition(alt_filename, patterns)
            if not led_exists:
                e_ref, e_sp, e_wp, e_t = extract_first_match_from_file(alt_filename, patterns, method, use_ref_as_rhf_in_hfld)
                ref_matrix = np.zeros((matrix_size, matrix_size))
                sp_matrix = np.zeros((matrix_size, matrix_size))
                wp_matrix = np.zeros((matrix_size, matrix_size))
                t_matrix = np.zeros((matrix_size, matrix_size))

                # Place values on the first diagonal element
                ref_matrix[0, 0] = e_ref
                sp_matrix[0, 0] = e_sp
                wp_matrix[0, 0] = e_wp
                t_matrix[0, 0] = e_t

                # Use the system label and add ' ALT' to the prefix
                file_prefix = system_label + ' ALT'
                matrices[f'Intra REF {file_prefix}'] = ref_matrix
                matrices[f'Intra SP {file_prefix}'] = sp_matrix
                matrices[f'Intra WP {file_prefix}'] = wp_matrix
                matrices[f'Intra T {file_prefix}'] = t_matrix

    # Ensure the output directory exists
    os.makedirs(LEDAW_output_path, exist_ok=True)

    # Define the full path for the output file
    output_file = os.path.join(LEDAW_output_path, 'tmp2.xlsx')

    # Writing matrices to the specified output file
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        for name, matrix in matrices.items():
            df = pd.DataFrame(matrix, index=range(1, matrix_size + 1), columns=range(1, matrix_size + 1))
            df.to_excel(writer, sheet_name=name)

    return matrices


def collect_unprocessed_LED_data_as_matrices(LEDAW_output_path):
    """Combines two temporary Excel files with multiple sheets into one file and deletes the source Excel files."""
    file1 = os.path.join(LEDAW_output_path, "tmp1.xlsx")
    file2 = os.path.join(LEDAW_output_path, "tmp2.xlsx")
    write_to_excel_filename = os.path.join(LEDAW_output_path, "Unprocessed_LED_matrices.xlsx")
    
    # Create a new Excel writer object
    with pd.ExcelWriter(write_to_excel_filename) as writer:
        # Load the first file and write its sheets to the output file
        with pd.ExcelFile(file1) as xls1:
            for sheet_name in xls1.sheet_names:
                df = pd.read_excel(xls1, sheet_name=sheet_name)
                # If the first column has the 'Unnamed' pattern, remove it, but only if the DataFrame is not empty
                if not df.empty and df.columns[0] == "Unnamed: 0":
                    df = df.rename(columns={"Unnamed: 0": ""})
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Load the second file and write its sheets to the output file
        with pd.ExcelFile(file2) as xls2:
            for sheet_name in xls2.sheet_names:
                df = pd.read_excel(xls2, sheet_name=sheet_name)
                # If the first column has the 'Unnamed' pattern, remove it, but only if the DataFrame is not empty
                if not df.empty and df.columns[0] == "Unnamed: 0":
                    df = df.rename(columns={"Unnamed: 0": ""})
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Delete the source Excel files
    os.remove(file1)
    os.remove(file2)
    print(f"LED energy components from each file were written as matrices without any processing to {write_to_excel_filename}")
	

def reorder_labels(system_labels, alternative_labels, main_label_mappings, alternative_label_mappings, LEDAW_output_path):
    """Process each sheet in the Excel file and reorder matrices based on dynamic prefix mappings."""
    
    # Define the input and output Excel file names
    input_excel_file = os.path.join(LEDAW_output_path, 'Unprocessed_LED_matrices.xlsx')
    base_name = os.path.splitext(input_excel_file)[0]
    output_excel_file = f"{base_name}_withFragmentRelabeling.xlsx"
    
    # Create mapping dictionaries for system labels and alternative labels
    main_mappings = dict(zip(system_labels, main_label_mappings))
    alternative_mappings = dict(zip(alternative_labels, alternative_label_mappings))
    
    xl = pd.ExcelFile(input_excel_file)
    
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        for sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name, index_col=0)
            
            is_alternative = 'ALT' in sheet_name
            label_prefix = None
            
            # Determine the correct label prefix from the sheet name
            for label in system_labels + alternative_labels:
                if label in sheet_name:
                    label_prefix = label
                    break
            
            if label_prefix:
                # Choose the appropriate mapping based on whether the sheet is from a main or alternative label
                if is_alternative and label_prefix in alternative_labels:
                    label_mapping = alternative_mappings[label_prefix]
                elif label_prefix in system_labels:
                    label_mapping = main_mappings[label_prefix]
                else:
                    raise ValueError(f"Unexpected label {label_prefix} in sheet {sheet_name}")
                
                # Apply the label mapping to reorder the DataFrame
                df.columns = label_mapping
                df.index = label_mapping

                # Sort the DataFrame according to the new mapping
                df.sort_index(axis=0, ascending=True, inplace=True)
                df.sort_index(axis=1, ascending=True, inplace=True)

                # Write the reordered DataFrame to the new Excel file
                df.to_excel(writer, sheet_name=sheet_name)
            else:
                if sheet_name != 'Sheet1':
                    print(f"Warning: No matching prefix found for sheet {sheet_name}. Sheet will not be reordered.")
                
                # Optionally write the sheet without reordering
                df.to_excel(writer, sheet_name=sheet_name)

    print(f"Reordered matrices were written to '{output_excel_file}'")



def compare_main_ALT_removeALTlabel(LEDAW_output_path):
    """Process the Excel file to compare and handle ALT sheets, then write results to a new file."""
    
    # Static file names with path
    input_excel_file = os.path.join(LEDAW_output_path, 'Unprocessed_LED_matrices_withFragmentRelabeling.xlsx')
    output_excel_file = os.path.join(LEDAW_output_path, 'Unprocessed_LED_matrices_withoutALTlabel.xlsx')
    
    difference_detected = False  # Flag to check if any differences were found
    differences = []  # List to store sheet names with differences
    xl = pd.ExcelFile(input_excel_file)
    sheets_to_remove = []
    sheets_to_rename = {}

    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name, index_col=0)

        # Identify corresponding ALT sheets
        if 'ALT' in sheet_name:
            original_sheet_name = sheet_name.replace(' ALT', '')
            if original_sheet_name in xl.sheet_names:
                df_original = xl.parse(original_sheet_name, index_col=0)
                
                if df_original.values.sum() == 0:
                    sheets_to_remove.append(original_sheet_name)
                    sheets_to_rename[sheet_name] = original_sheet_name
                elif df.values.sum() == 0:
                    sheets_to_remove.append(sheet_name)
                else:
                    diff = abs(df - df_original)
                    if (diff.values > 1e-4).any():
                        differences.append(original_sheet_name)
                    else:
                        sheets_to_remove.append(sheet_name)
            else:
                sheets_to_rename[sheet_name] = original_sheet_name

    # Write the processed sheets to a new Excel file
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        for sheet_name in xl.sheet_names:
            if sheet_name not in sheets_to_remove:
                df = xl.parse(sheet_name, index_col=0)
                if sheet_name in sheets_to_rename:
                    sheet_name = sheets_to_rename[sheet_name]
                df.to_excel(writer, sheet_name=sheet_name)
    
    # Print message before raising the exception
    print(f"Redundant matrices from main and alternative files were eliminated. Cleaned data were written to '{output_excel_file}'")

    # Handle differences after writing to the new Excel file
    if differences:
        for sheet_name in differences:
            print(f"Matrices in '{sheet_name}' and its alternative are different")
        # Raise the exception after printing the message
        raise MatrixDifferenceError("Significant differences were found in the above listed matrices. Execution was continued with the data in the main file.")


def combine_intra_inter_matrices(LEDAW_output_path):
    """Combines Intra and Inter matrices from an Excel file and writes the results to a new file."""
    
    # Static file names with paths
    input_excel_file = os.path.join(LEDAW_output_path, 'Unprocessed_LED_matrices_withoutALTlabel.xlsx')
    output_excel_file = os.path.join(LEDAW_output_path, 'Unprocessed_LED_matrices_combined_INTRA-INTER.xlsx')
    
    # Load the input Excel file
    xl = pd.ExcelFile(input_excel_file)
    combined_sheets = {}
    sheets_to_transfer = []

    # Group sheets by their base names without "Intra" or "Inter"
    sheet_groups = {}
    for sheet_name in xl.sheet_names:
        if 'Intra' in sheet_name:
            base_name = sheet_name.replace('Intra ', '')
            if base_name not in sheet_groups:
                sheet_groups[base_name] = {'Intra': None, 'Inter': None}
            sheet_groups[base_name]['Intra'] = sheet_name
        elif 'Inter' in sheet_name:
            base_name = sheet_name.replace('Inter ', '')
            if base_name not in sheet_groups:
                sheet_groups[base_name] = {'Intra': None, 'Inter': None}
            sheet_groups[base_name]['Inter'] = sheet_name
        else:
            # Non-Intra/Inter sheets are directly added for transfer
            sheets_to_transfer.append(sheet_name)

    # Combine Intra and Inter sheets
    for base_name, sheets in sheet_groups.items():
        intra_df = xl.parse(sheets['Intra'], index_col=0) if sheets['Intra'] else None
        inter_df = xl.parse(sheets['Inter'], index_col=0) if sheets['Inter'] else None

        if intra_df is not None and inter_df is not None:
            combined_df = intra_df + inter_df
            combined_sheets[base_name] = combined_df
        elif intra_df is not None:
            combined_sheets[base_name] = intra_df
        elif inter_df is not None:
            combined_sheets[base_name] = inter_df

    # Write the combined matrices and other sheets to the output Excel file
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        # Write combined intra-inter matrices
        for base_name, combined_df in combined_sheets.items():
            combined_df.to_excel(writer, sheet_name=base_name)
        
        # Transfer the remaining sheets
        for sheet_name in sheets_to_transfer:
            df = xl.parse(sheet_name, index_col=0)
            df.to_excel(writer, sheet_name=sheet_name)
    
    print(f"Intra and Inter matrices were combined and written to '{output_excel_file}'")


def compute_all_standard_led_int_en_matrices(system_labels, conversion_factor, method, LEDAW_output_path, main_subsystem_matching_labels, main_filenames, alternative_filenames):
    """Process LED matrices of supersystem and its subsystems to compute LED interaction energy map and its components, and write results to a new Excel file."""
    
    # Define file paths with the provided LEDAW_output_path
    input_excel_file = os.path.join(LEDAW_output_path, 'Unprocessed_LED_matrices_combined_INTRA-INTER.xlsx')
    output_excel_file = os.path.join(LEDAW_output_path, 'Unrelabeled_All_Standard_LED_matrices.xlsx')
    
    supersystem_label = system_labels[0]
    subsystem_labels = system_labels[1:]

    matrices = {}
    xl = pd.ExcelFile(input_excel_file)

    for sheet_name in xl.sheet_names:
        if supersystem_label in sheet_name:
            new_sheet_name = sheet_name.replace(supersystem_label, '').strip()
            df_supersystem = xl.parse(sheet_name, index_col=0)

            df_sum_subsystems = pd.DataFrame(0, index=df_supersystem.index, columns=df_supersystem.columns)

            for subsystem_label in subsystem_labels:
                subsystem_sheet_name = sheet_name.replace(supersystem_label, subsystem_label)
                if subsystem_sheet_name in xl.sheet_names:
                    df_subsystem = xl.parse(subsystem_sheet_name, index_col=0)
                    
                    # Handle duplicate indices by keeping the first occurrence and dropping the rest
                    df_subsystem = df_subsystem[~df_subsystem.index.duplicated(keep='first')]
                    df_subsystem = df_subsystem.loc[:, ~df_subsystem.columns.duplicated(keep='first')]

                    df_sum_subsystems += df_subsystem

            # Apply the conversion factor or compute the difference depending on the method
            if method.lower() == "hfld" and ('Disp SP' in sheet_name or 'WP' in sheet_name or 'Disp HFLD' in sheet_name):
                for i, sublist in enumerate(main_subsystem_matching_labels):
                    for j in range(len(sublist)):
                        for k in range(j + 1, len(sublist)):
                            df_supersystem.iat[sublist[j] - 1, sublist[k] - 1] = 0
                            df_supersystem.iat[sublist[k] - 1, sublist[j] - 1] = 0

                df_result = df_supersystem * conversion_factor  # Use only supersystem values for dispersion in HFLD
            else:
                df_result = (df_supersystem - df_sum_subsystems) * conversion_factor  # Compute differences for other matrices

            df_result = df_result.where(np.triu(np.ones(df_result.shape), k=0).astype(bool))
            matrices[new_sheet_name] = df_result

    # Handle Dispersion matrices for DLPNO-CCSD(T), DLPNO-CCSD, and HFLD methods
    if 'Disp SP' in matrices:
        df_disp_wp = matrices.get('WP', pd.DataFrame(np.nan, index=matrices['SP'].index, columns=matrices['SP'].columns)).copy()
        np.fill_diagonal(df_disp_wp.values, np.nan)
        matrices['Disp WP'] = df_disp_wp

        df_disp_sp = matrices['Disp SP']
        df_disp_ccsd = df_disp_sp + df_disp_wp
        
        if method.lower() == "hfld":
            matrices['Disp HFLD'] = df_disp_ccsd
            matrices['C-HFLD'] = df_disp_ccsd

    if method.lower() == "dlpno-ccsd(t)" and 'T' in matrices:
        df_disp_t = matrices['T']
        df_disp_sp = matrices['Disp SP']
        df_disp_wp = matrices['Disp WP']
        
        with np.errstate(divide='ignore', invalid='ignore'):
            df_disp_t = np.divide(df_disp_sp * df_disp_t, matrices['SP'])
            df_disp_t[np.isnan(df_disp_t)] = 0

        df_disp_t = df_disp_t.where(np.triu(np.ones(df_disp_t.shape), k=0).astype(bool))
        matrices['Disp T'] = df_disp_t
        
        df_disp_ccsd_t = df_disp_sp + df_disp_wp + df_disp_t
        matrices['Disp CCSD(T)'] = df_disp_ccsd_t

        if 'Singles' in matrices:
            df_singles = matrices['Singles']
            df_c_ccsd_t = matrices['SP'] + matrices['WP'] + matrices['T'] + df_singles
            matrices['C-CCSD(T)'] = df_c_ccsd_t
        else:
            df_c_ccsd_t = matrices['SP'] + matrices['WP'] + matrices['T']
            matrices['C-CCSD(T)'] = df_c_ccsd_t

        df_inter_non_disp_ccsd_t = matrices['C-CCSD(T)'] - matrices['Disp CCSD(T)']
        matrices['Inter-NonDisp-C-CCSD(T)'] = df_inter_non_disp_ccsd_t

        if 'REF' in matrices:
            df_total = matrices['REF'] + matrices['C-CCSD(T)']
            matrices['TOTAL'] = df_total

    if method.lower() == "dlpno-ccsd" and 'Singles' in matrices:
        df_singles = matrices['Singles']
        df_c_ccsd = matrices['SP'] + matrices['WP'] + df_singles
        matrices['C-CCSD'] = df_c_ccsd

        df_inter_non_disp_ccsd = matrices['C-CCSD'] - matrices['Disp CCSD']
        matrices['Inter-NonDisp-C-CCSD'] = df_inter_non_disp_ccsd

        if 'REF' in matrices:
            df_total = matrices['REF'] + matrices['C-CCSD']
            matrices['TOTAL'] = df_total

    if method.lower() == "hfld" and 'C-HFLD' in matrices:
        if 'REF' in matrices:
            df_total = matrices['REF'].copy()
            df_c_hfld = matrices['C-HFLD']
            
            for i in range(df_c_hfld.shape[0]):
                for j in range(df_c_hfld.shape[1]):
                    if not np.isnan(df_c_hfld.iat[i, j]):
                        df_total.iat[i, j] += df_c_hfld.iat[i, j]
            
            matrices['TOTAL'] = df_total

    # Compute the dielectric interaction energy using compute_diel_int_en
    diel_values, diel_int_energy = compute_diel_int_en(main_filenames, alternative_filenames, conversion_factor)

    # Add the DIEL matrix and update TOTAL
    if 'TOTAL' in matrices and diel_int_energy != 0:
        total_matrix = matrices['TOTAL']

        # Calculate total sum without DIEL (sum of all elements)
        total_sum_wo_diel = total_matrix.sum().sum()

        if total_sum_wo_diel != 0:
            diel_matrix = total_matrix * (diel_int_energy / total_sum_wo_diel)
            matrices['DIEL'] = diel_matrix
            matrices['TOTAL'] = total_matrix + diel_matrix  # Add DIEL to TOTAL

    # Write the matrices from the dictionary to the output Excel file
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        # Write TOTAL and DIEL if they exist
        if 'TOTAL' in matrices:
            matrices['TOTAL'].to_excel(writer, sheet_name='TOTAL')
        if 'DIEL' in matrices:
            matrices['DIEL'].to_excel(writer, sheet_name='DIEL')

        # Write all other matrices
        for sheet_name, df in matrices.items():
            if sheet_name not in ['TOTAL', 'DIEL']:  # Avoid writing TOTAL twice
                df.to_excel(writer, sheet_name=sheet_name)

    print(f"All standard '{method}/LED' interaction energy matrices were written to '{output_excel_file}'")


def relabel_and_sort_matrices(relabel_mapping, LEDAW_output_path, method):
    """Relabels and sorts the LED matrices to change fragment labels."""
    
    # Define file paths with the provided LEDAW_output_path
    input_excel_file = os.path.join(LEDAW_output_path, 'Unrelabeled_All_Standard_LED_matrices.xlsx')
    output_excel_file = os.path.join(LEDAW_output_path, 'All_Standard_LED_matrices.xlsx')
    
    # Check if relabel_mapping is defined and not empty
    if not relabel_mapping:
        # If relabel_mapping is empty or not provided, rename the input file to the output file without any changes
        shutil.copy(input_excel_file, output_excel_file)
        print(f"No relabeling applied. The file was copied as '{output_excel_file}' without changes.")
        return
    
    # Load the Excel file
    xl = pd.ExcelFile(input_excel_file)

    # Create a new Excel writer object to save the modified matrices
    with pd.ExcelWriter(output_excel_file, engine='openpyxl', mode='w') as writer:
        for sheet_name in xl.sheet_names:
            df = xl.parse(sheet_name, index_col=0)

            # Apply the relabeling
            df.columns = [df.columns[i-1] for i in relabel_mapping]
            df.index = [df.index[i-1] for i in relabel_mapping]

            # Sort the DataFrame
            df.sort_index(axis=0, ascending=True, inplace=True)
            df.sort_index(axis=1, ascending=True, inplace=True)

            # Ensure symmetry in the DataFrame
            for i in range(1, df.shape[0]):
                for j in range(i):
                    # Swap the corresponding elements to ensure symmetry
                    if np.isnan(df.iloc[j, i]) and not np.isnan(df.iloc[i, j]):
                        df.iloc[j, i] = df.iloc[i, j]
                        df.iloc[i, j] = np.nan

            # Write the relabeled DataFrame back to the Excel file
            df.to_excel(writer, sheet_name=sheet_name)

    print(f"All standard '{method}/LED' interaction energy matrices after relabeling and sorting fragments were written to '{output_excel_file}'")


def write_standard_LED_summary_int_en_matrices(method, LEDAW_output_path):
    """Write summary standard LED interaction energy maps to an excel file"""
    
    # Define the input and output Excel file names with path handling
    input_excel_file = os.path.join(LEDAW_output_path, 'All_Standard_LED_matrices.xlsx')
    output_excel_file = os.path.join(LEDAW_output_path, 'Summary_Standard_LED_matrices.xlsx')
    
    # Define the sheets to be written
    if method.lower() == "dlpno-ccsd(t)":
        sheets_to_transfer = [
            'TOTAL', 'DIEL', 'REF', 'Electrostat', 'Exchange', 
            'C-CCSD(T)', 'Disp CCSD(T)', 'Inter-NonDisp-C-CCSD(T)'
        ]
    elif method.lower() == "dlpno-ccsd":
        sheets_to_transfer = [
            'TOTAL', 'DIEL', 'REF', 'Electrostat', 'Exchange', 
            'C-CCSD', 'Disp CCSD', 'Inter-NonDisp-C-CCSD'
        ]
    elif method.lower() == "hfld":
        sheets_to_transfer = [
            'TOTAL', 'DIEL', 'REF', 'Electrostat', 'Exchange', 
            'Disp HFLD'
        ]
    else:
        raise ValueError(f"Unsupported method: {method}. Please specify 'DLPNO-CCSD(T)', 'DLPNO-CCSD', or 'HFLD'.")

    # Load the workbook
    wb_input = load_workbook(input_excel_file)
    wb_output = Workbook()
    wb_output.remove(wb_output.active)

    for sheet_name in sheets_to_transfer:
        if sheet_name in wb_input.sheetnames:
            # Copy the sheet as is
            sheet = wb_input[sheet_name]
            wb_output.create_sheet(sheet_name)
            new_sheet = wb_output[sheet_name]

            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)
        else:
            print(f"Warning: Sheet '{sheet_name}' not found in {input_excel_file}.")

    # Save the output workbook
    wb_output.save(output_excel_file)
    print(f"The summary standard '{method}/LED' matrices were written to '{output_excel_file}'")


def compute_fp_el_prep(df_ref):
    """Compute fragment pairwise electronic preparation matrices."""
    diagonal_elements = np.diag(df_ref.values)
    distributed_matrix = np.zeros(df_ref.shape)

    non_diagonal_sums = np.array([compute_denominator_for_fp_el_prep(df_ref, i) for i in range(df_ref.shape[0])])

    for i in range(df_ref.shape[0]):
        for j in range(i + 1, df_ref.shape[1]):  # Only consider the upper triangle (i < j)
            if non_diagonal_sums[i] != 0 and non_diagonal_sums[j] != 0:
                term_1 = (diagonal_elements[i] * df_ref.iloc[i, j]) / non_diagonal_sums[i]
                term_2 = (diagonal_elements[j] * df_ref.iloc[i, j]) / non_diagonal_sums[j]
                distributed_value = term_1 + term_2
                distributed_matrix[i, j] = distributed_value

    # Convert to DataFrame and set diagonal and below-diagonal elements to NaN
    distributed_df = pd.DataFrame(distributed_matrix, index=df_ref.index, columns=df_ref.columns)
    np.fill_diagonal(distributed_df.values, np.nan)
    distributed_df = distributed_df.where(np.triu(np.ones(distributed_df.shape), k=1).astype(bool))

    return distributed_df


def compute_denominator_for_fp_el_prep(df, index):
    """Compute the sum of non-diagonal elements involving a particular index (both row and column)."""
    row_sum = df.iloc[index, :].sum() - df.iloc[index, index]
    col_sum = df.iloc[:, index].sum() - df.iloc[index, index]
    return row_sum + col_sum


def process_fp_LED_matrices(method, main_filenames, alternative_filenames, conversion_factor, LEDAW_output_path):
    """Process the matrices as per the given method and write to the output Excel file."""
    input_excel_file = os.path.join(LEDAW_output_path, 'Summary_Standard_LED_matrices.xlsx')
    output_excel_file = os.path.join(LEDAW_output_path, 'Summary_fp-LED_matrices.xlsx')

    # Load the input Excel file
    xl = pd.ExcelFile(input_excel_file)

    # Process REF sheet
    df_ref = pd.read_excel(xl, sheet_name='REF', index_col=0)
    df_ref_distributed = compute_fp_el_prep(df_ref)
    df_ref_distributed.columns.name = 'REF-EL-PREP'

    # Call the compute_diel_int_en function to get dielectric values and calculate dielectric interaction energy
    diel_values, e_diel_int_en = compute_diel_int_en(main_filenames, alternative_filenames, conversion_factor)

    # Initialize a new Excel writer object
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:

        # Load Electrostat and Exchange sheets for all methods
        df_electrostat = pd.read_excel(xl, sheet_name='Electrostat', index_col=0)
        df_exchange = pd.read_excel(xl, sheet_name='Exchange', index_col=0)

        if method.lower() in ["dlpno-ccsd(t)", "dlpno-ccsd"]:
            df_total = pd.read_excel(xl, sheet_name='TOTAL', index_col=0)
            df_total_distributed = compute_fp_el_prep(df_total)
            df_total_distributed.columns.name = 'C-CCSD(T)-EL-PREP' if method.lower() == "dlpno-ccsd(t)" else 'C-CCSD-EL-PREP'

            # Calculate TOTAL - REF and write it as EL-PREP-C-CCSD or EL-PREP-C-CCSD(T)
            df_c_ccsd = df_total_distributed - df_ref_distributed
            df_c_ccsd.columns.name = 'CCSD(T)-EL-PREP' if method.lower() == "dlpno-ccsd(t)" else 'CCSD-EL-PREP'

            # Calculate the final TOTAL as C-CCSD(T) + REF
            df_disp_ccsd = pd.read_excel(xl, sheet_name='Disp CCSD' if method.lower() == "dlpno-ccsd" else 'Disp CCSD(T)', index_col=0)
            df_inter_nondisp = pd.read_excel(xl, sheet_name='Inter-NonDisp-C-CCSD' if method.lower() == "dlpno-ccsd" else 'Inter-NonDisp-C-CCSD(T)', index_col=0)

            df_ref_final = df_electrostat + df_exchange + df_ref_distributed
            df_c_ccsd_final = df_c_ccsd + df_disp_ccsd + df_inter_nondisp
            df_total_final = df_ref_final + df_c_ccsd_final

            # If dielectric interaction energy is not zero, calculate and add the DIEL matrix
            if e_diel_int_en != 0:
                total_sum = df_total_final.sum().sum()
                diel_matrix = df_total_final * (e_diel_int_en / total_sum)
                df_diel_final = diel_matrix

                # Add DIEL to the TOTAL matrix
                df_total_final += df_diel_final

                # Write DIEL as the next sheet after TOTAL
                df_total_final.to_excel(writer, sheet_name='TOTAL')
                df_diel_final.to_excel(writer, sheet_name='DIEL')

            else:
                # Write the final TOTAL to the output file
                df_total_final.to_excel(writer, sheet_name='TOTAL')

            # Continue writing the rest of the sheets
            df_ref_final.to_excel(writer, sheet_name='REF')
            df_electrostat.to_excel(writer, sheet_name='Electrostat')
            df_exchange.to_excel(writer, sheet_name='Exchange')
            df_ref_distributed.to_excel(writer, sheet_name='REF-EL-PREP')
            df_c_ccsd_final.to_excel(writer, sheet_name='C-CCSD(T)' if method.lower() == "dlpno-ccsd(t)" else 'C-CCSD')
            df_disp_ccsd.to_excel(writer, sheet_name='Disp CCSD(T)' if method.lower() == "dlpno-ccsd(t)" else 'Disp CCSD')
            df_inter_nondisp.to_excel(writer, sheet_name='Inter-NonDisp-C-CCSD(T)' if method.lower() == "dlpno-ccsd(t)" else 'Inter-NonDisp-C-CCSD')
            df_c_ccsd.to_excel(writer, sheet_name='C-CCSD(T)-EL-PREP' if method.lower() == "dlpno-ccsd(t)" else 'C-CCSD-EL-PREP')
            df_total_distributed.to_excel(writer, sheet_name='CCSD(T)-EL-PREP' if method.lower() == "dlpno-ccsd(t)" else 'CCSD-EL-PREP')

        elif method.lower() == "hfld":
            # Calculate REF = Electrostat + Exchange + REF-EL-PREP
            df_ref_final = df_electrostat + df_exchange + df_ref_distributed

            # Calculate TOTAL = REF + Disp HFLD
            df_disp_hfld = pd.read_excel(xl, sheet_name='Disp HFLD', index_col=0)
            df_total_hfld = df_ref_final + df_disp_hfld
            df_total_hfld.columns.name = 'TOTAL'

            # If dielectric interaction energy is not zero, calculate and add the DIEL matrix
            if e_diel_int_en != 0:
                total_sum = df_total_hfld.sum().sum()
                diel_matrix = df_total_hfld * (e_diel_int_en / total_sum)
                df_diel_final = diel_matrix

                # Add DIEL to the TOTAL matrix
                df_total_hfld += df_diel_final

                # Write DIEL as the next sheet after TOTAL
                df_total_hfld.to_excel(writer, sheet_name='TOTAL')
                df_diel_final.to_excel(writer, sheet_name='DIEL')
            else:
                # Write the fp-TOTAL sheet without DIEL
                df_total_hfld.to_excel(writer, sheet_name='TOTAL')

            # Continue writing the rest of the sheets
            df_ref_final.to_excel(writer, sheet_name='REF')
            df_electrostat.to_excel(writer, sheet_name='Electrostat')
            df_exchange.to_excel(writer, sheet_name='Exchange')
            df_ref_distributed.to_excel(writer, sheet_name='REF-EL-PREP')
            df_disp_hfld.to_excel(writer, sheet_name='Disp HFLD')

        print(f"fp-LED interaction energy matrices were written to '{output_excel_file}'")


def delete_unprocessed_files(LEDAW_output_path):
    """
    Deletes all files in the specified directory that start with 'Un'.
    """

    # Construct the pattern for matching files starting with 'Un'
    pattern = os.path.join(LEDAW_output_path, 'Un*')
    
    # Find all files that match the pattern
    files_to_delete = glob.glob(pattern)
    
    # Delete each file found
    print("Temporary files written will be deleted:")
    for file_path in files_to_delete:
        try:
            os.remove(file_path)
            print(f"Deleted: {file_path}")
        except Exception as e:
            print(f"Error deleting {file_path}: {e}")


def engine_LED_N_body(main_filenames, alternative_filenames, conversion_factor, method, LEDAW_output_path, relabel_mapping=None, use_ref_as_rhf_in_hfld=None):
    """
    Engine function to process LED N-body interaction energy matrices, standardize and reorder labels,
    compute final matrices, and provide summaries.
    """

    # Ensure alternative filenames are valid or empty
    alternative_filenames = [filename if filename else None for filename in alternative_filenames]    

    # Label main filenames
    labeled_main_filenames = label_systems(main_filenames)

    # Use main labels for alt filenames where the alt filenames are None
    labeled_alt_filenames = {}
    for main_file, alt_file in zip(main_filenames, alternative_filenames):
        if alt_file:
            labeled_alt_filenames[alt_file] = labeled_main_filenames[main_file]
        else:
            labeled_alt_filenames[main_file] = labeled_main_filenames[main_file]  # Use main label if alt is None

    # Construct label mapping list wrt. supersystem labels to standardize labeling between super and subsystem files
    main_label_mappings, match_dicts_main, subsystem_matching_labels_main = construct_label_mappings(main_or_alt_filenames=list(labeled_main_filenames.keys()))
    alternative_label_mappings, match_dicts_alt, subsystem_matching_labels_alt = construct_label_mappings(main_or_alt_filenames=list(labeled_alt_filenames.keys()), default_label_mappings=main_label_mappings)

    # Call multifrag_system_processing and singlefrag_system_processing and combine their files
    matrix_size = multifrag_system_processing(main_filenames=list(labeled_main_filenames.keys()), 
                                             alternative_filenames=list(labeled_alt_filenames.keys()), 
                                             LEDAW_output_path=LEDAW_output_path,
                                             system_labels=list(labeled_main_filenames.values()))

    matrices = singlefrag_system_processing(main_filenames=list(labeled_main_filenames.keys()), 
        alternative_filenames=list(labeled_alt_filenames.keys()), 
        matrix_size=matrix_size, LEDAW_output_path=LEDAW_output_path,
        system_labels=list(labeled_main_filenames.values()),
        method=method,
        use_ref_as_rhf_in_hfld=use_ref_as_rhf_in_hfld)

    collect_unprocessed_LED_data_as_matrices(LEDAW_output_path=LEDAW_output_path)

    # Reorder the labels
    reorder_labels(system_labels=list(labeled_main_filenames.values()), 
        alternative_labels=list(labeled_alt_filenames.values()), 
        main_label_mappings=main_label_mappings, 
        alternative_label_mappings=alternative_label_mappings, 
        LEDAW_output_path=LEDAW_output_path)

    # Get rid of ALT labels
    try:
        compare_main_ALT_removeALTlabel(LEDAW_output_path=LEDAW_output_path)
    except MatrixDifferenceError as e:
        print(f"PLEASE CHECK THE INCONSISTENCY IN YOUR DATA: {e}")
        return

    # Combine intra and inter matrices
    combine_intra_inter_matrices(LEDAW_output_path=LEDAW_output_path)

    compute_all_standard_led_int_en_matrices(system_labels=list(labeled_main_filenames.values()),
                                             conversion_factor=conversion_factor, 
                                             method=method, LEDAW_output_path=LEDAW_output_path,
                                             main_subsystem_matching_labels=subsystem_matching_labels_main,
                                             main_filenames=main_filenames,
                                             alternative_filenames=alternative_filenames)

    # If relabeling is not specified, determine the number of fragments based on the supersystem labels 
    # and construct the original label list
    if relabel_mapping is None or relabel_mapping == []:
        max_label = max(main_label_mappings[0])
        relabel_mapping = list(range(1, max_label + 1))

    # Relabel and sort the interaction energy matrices if relabel_mapping is specified
    relabel_and_sort_matrices(relabel_mapping=relabel_mapping, LEDAW_output_path=LEDAW_output_path, method=method)

    # Write summary standard LED interaction energy matrices to an excel File
    write_standard_LED_summary_int_en_matrices(method=method, 
                                               LEDAW_output_path=LEDAW_output_path)

    # Write fp-LED Interaction Energy Matrices to an Excel File
    process_fp_LED_matrices(method=method, 
                            main_filenames=list(labeled_main_filenames.keys()), 
                            alternative_filenames=list(labeled_alt_filenames.keys()), 
                            conversion_factor=conversion_factor, 
                            LEDAW_output_path=LEDAW_output_path)

    # Delete temporary unprocessed files
    delete_unprocessed_files(LEDAW_output_path=LEDAW_output_path)
    
    print('\n')
    print('*'*120)
    print(f"  N-body LED analyses were terminated NORMALLY. Standard and fp-LED matrices are at {LEDAW_output_path}")
    print('*'*120)
